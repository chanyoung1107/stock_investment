# -*- coding: utf-8 -*-
from PyQt5 import QtWidgets
from PyQt5.QtCore import *
import win32com.client
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from openpyxl import Workbook
import sys
import datetime
import operator


cybos = ''

layout = ''
emptyLabel = ''
loginCheckBtn = ''
loginStatusLabel = ''
loginYNStatusLabel = ''
status = False
startBtn = ''

processStatusLabel = ''

perList = []
fullDataDictList = []
fullDataList = []
dataDict = {}
dataList = []
per_roa_dict = {}
per_dict = {}
roa_dict = {}


class StockStart(QWidget):
    def __init__(self):
        super().__init__()
        self.appStart()

    def appStart(self):
        global layout
        global emptyLabel
        global loginCheckBtn
        global loginStatusLabel
        global startBtn
        global loginYNStatusLabel
        global processStatusLabel

        # 위젯 속성 지정
        self.setWindowTitle('DAISHIN STOCK')
        self.setFixedSize(500, 250)
        self.setFocusPolicy(Qt.StrongFocus)

        # 프로그램 시작 레이아웃
        layout = QtWidgets.QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # 레이아웃 설정하기
        self.setLayout(layout)

        # 프로그램 시작시 나타나는 버튼 생성
        emptyLabel = QtWidgets.QLabel()
        loginCheckBtn = QtWidgets.QPushButton("Login Check")
        loginStatusLabel = QtWidgets.QLabel()
        startBtn = QtWidgets.QPushButton("Start")
        loginYNStatusLabel = QtWidgets.QLabel()
        processStatusLabel = QtWidgets.QLabel()

        layout.addWidget(emptyLabel)
        layout.addWidget(loginCheckBtn)
        layout.addWidget(loginStatusLabel)
        layout.addWidget(startBtn)
        layout.addWidget(loginYNStatusLabel)

        loginCheckBtn.clicked.connect(self.loginCheck)
        startBtn.clicked.connect(self.startStock)

    def loginCheck(self):
        global cybos
        global loginCheckBtn
        global loginStatusLabel
        global startBtn
        global status

        # CpCybos - CYBOS의 각종 상태를 확인할 수 있음. (모듈 위치: CpUtil.dll)
        cybos = win32com.client.Dispatch("CpUtil.CpCybos")
        # print(cybos.IsConnect)           # 연결상태 확인
        if cybos.IsConnect != 1:
            loginStatusLabel.setText("사이보스에 연결되지 않았습니다. 관리자모드로 실행하여 주세요.")
            loginStatusLabel.setStyleSheet("font-weight: bold;")
            status = False
        else:
            loginStatusLabel.setText("연결상태 : 정상")
            loginStatusLabel.setStyleSheet("font-weight: bold;")
            status = True

    def startStock(self):
        global status
        global loginYNStatusLabel

        if status:
            self.runningReady()
        else:
            loginYNStatusLabel.setText("Login Check 버튼을 눌러 정상 확인하여 주세요.")
            loginYNStatusLabel.setStyleSheet("font-weight: bold;")

    def runningReady(self):
        global layout
        global emptyLabel
        global loginCheckBtn
        global loginStatusLabel
        global startBtn
        global loginYNStatusLabel
        global status
        global processStatusLabel

        # 파일선택버튼 비활성화
        loginCheckBtn.setEnabled(False)
        startBtn.setEnabled(False)

        emptyLabel.hide()
        loginCheckBtn.hide()
        loginStatusLabel.hide()
        startBtn.hide()
        loginYNStatusLabel.hide()

        layout.addWidget(processStatusLabel)

        processStatusLabel.setText("프로그램이 가동중입니다.....")
        processStatusLabel.setStyleSheet("font-weight: bold;"
                                         "font-size: 25px;"
                                         "text-align: center;")

        self.processStart()

            
    def processStart(self):
        global processStatusLabel
        global perList
        global fullDataDictList
        global fullDataList
        global dataDict
        global dataList
        global per_roa_dict
        global per_dict
        global roa_dict

        write_xl = Workbook()
        write_ws = write_xl.active
        write_ws.merge_cells('A1:G1')
        date = str(datetime.datetime.now()).split('.')[0]
        weekend = '월화수목금토일'
        write_ws['A1'] = date + ' (' + weekend[datetime.datetime.now().weekday()] + ')'
        write_ws.merge_cells('A2:G2')
        write_ws['A2'] = 'PER 공식1 : 주가 / 주당순이익 - (주당순이익 = 당기순이익 / 주식수)'
        write_ws.merge_cells('A3:G3')
        write_ws['A3'] = 'PER 공식2 : 시가총액 / 당기순이익'
        write_ws.merge_cells('A4:G4')
        write_ws['A4'] = 'ROE 공식 : 당기순이익 / 자본총액'
        write_ws.merge_cells('A5:G5')
        write_ws['A5'] = '재무제표 비교 사이트: https://comp.fnguide.com'

        write_ws.append(['종목', 'PER', '', '종목', 'ROE (%)', '', '순위종목', '종목코드', '순위합산', '', 'PER', 'ROE (%)', '당기순이익 (억)', '전일종가', '총자산 (억)'])

        # 주식 종록에 대한 정보 확인
        cpStockCode = win32com.client.Dispatch("CpUtil.CpStockCode")
        # print(cpStockCode.GetCount())        # 주식 상장(비상장 일부 포함) 갯수
        # print(cpStockCode.GetData(1, 1))     # 주식 종목(0: 종목코드, 1: 종목명, 2:둘다) / 인자값
        # 여러 종목의 필요 항목을 한번에 수신
        marketEye = win32com.client.Dispatch("CpSysDib.MarketEye")
        # 업종별 코드 리스트
        cpCodeMgr = win32com.client.Dispatch("CpUtil.CpCodeMgr")

        # kospi / kosdac 종목코드 리스트
        kospiList = cpCodeMgr.getstocklistbymarket(1)
        kosdacList = cpCodeMgr.getstocklistbymarket(2)

        per_dict = {}
        roe_dict = {}

        for i in range(len(kospiList)):
            dataDict[cpCodeMgr.codetoname(kospiList[i])] = kospiList[i]
            dataList.append(kospiList[i])
            if i % 60 == 0 and i != 0:
                fullDataDictList.append(dataDict)
                fullDataList.append(dataList)
                dataDict = {}
                dataList = []
            if i == len(kospiList) - 1:
                fullDataDictList.append(dataDict)
                fullDataList.append(dataList)

        for i in range(len(fullDataList)):
            marketEye.SetInputValue(0, (20, 23, 67, 75, 77, 88, 89))
            marketEye.SetInputValue(1, fullDataList[i])
            marketEye.BlockRequest()

            for idx, (key, value) in enumerate(fullDataDictList[i].items()):
                if marketEye.getDataValue(2, idx) != 0.0:
                    own_value = marketEye.getDataValue(0, idx) * marketEye.getDataValue(6, idx)
                    debt = marketEye.getDataValue(3, idx) * own_value / 100
                    total_value = (own_value + debt) / 100000000

                    # 종목: ['종목코드', 'PER', 'ROE (%)', '당기순이익', '전일종가', '총자산']
                    per_roa_dict[key] = (value, marketEye.getDataValue(2, idx), marketEye.getDataValue(4, idx),
                                         (marketEye.getDataValue(5, idx) / 100000000), marketEye.getDataValue(1, idx),
                                         total_value)

                    per_dict[key] = marketEye.getDataValue(2, idx)
                    roe_dict[key] = marketEye.getDataValue(4, idx)

        sort_per_list = sorted(per_dict.items(), key=lambda item: item[1])
        sort_roe_list = sorted(roe_dict.items(), reverse=True,  key=lambda item: item[1])
        sort_per_dict = {}
        sort_roe_dict = {}
        per_rank = {}
        roe_rank = {}
        total_rank = {}

        # sorting한 per/roe list dict로 변경 {이름: per/roe}
        for i in range(len(sort_per_list)):
            sort_per_dict[sort_per_list[i][0]] = sort_per_list[i][1]
            sort_roe_dict[sort_roe_list[i][0]] = sort_roe_list[i][1]

        per_rank_count = 1
        roe_rank_count = 1

        # per/roe dict 순위로 dict 추가 {이름: 순위}
        for key in sort_per_dict:
            per_rank[key] = per_rank_count
            per_rank_count += 1

        for key in sort_roe_dict:
            roe_rank[key] = roe_rank_count
            roe_rank_count += 1

        # total 순위합산 total_rank dict에 추가 {이름: 순위합산}
        for key, value in per_rank.items():
            for subKey, subValue in roe_rank.items():
                if key == subKey:
                    rank = per_rank[key] + roe_rank[subKey]
                    total_rank[key] = rank

        # print(per_roa_dict) # {'동화약품': ('A000020', 16.450000762939453, 3.140000104904175, 93.93, 16550, 4069.313113178766),}
        # 종목: ['종목코드', 'PER', 'ROE (%)', '당기순이익', '전일종가', '총자산']
        # print(per_rank) # '태영건설': 1, '미래아이앤지': 2, 'SK가스': 3, '한진중공업': 4, '원림': 5,
        # print(roe_rank) # '세하': 1, '진양폴리': 2, '범양건영': 3, 'KG케미칼': 4, '신풍제지': 5, '코웨이': 6,
        # print(total_rank)   # '태영건설': 174, '미래아이앤지': 25, 'SK가스': 142, '한진중공업': 482,
        # {이름: 순위합산} sorting
        sort_total_list = sorted(total_rank.items(), key=lambda item: item[1])

        # ['종목', '종목코드', 'PER', '', '종목', '종목코드', 'ROE (%)', '', '순위종목', '종목코드', '순위합산', '', 'PER', 'ROE (%)', '당기순이익 (억)', '전일종가', '총자산 (억)']
        for i in range(len(sort_total_list)):
            write_ws.append([sort_per_list[i][0], per_roa_dict[sort_per_list[i][0]][1], '',
                             sort_roe_list[i][0], per_roa_dict[sort_roe_list[i][0]][2], '',
                             sort_total_list[i][0], per_roa_dict[sort_total_list[i][0]][0], sort_total_list[i][1], '',
                             per_roa_dict[sort_total_list[i][0]][1], per_roa_dict[sort_total_list[i][0]][2], per_roa_dict[sort_total_list[i][0]][3],
                             per_roa_dict[sort_total_list[i][0]][4], per_roa_dict[sort_total_list[i][0]][5]])

        excelUrl = QFileDialog.getSaveFileName(self, 'Save xlsx file', filter="*.xlsx")  # 파일 경로 + 이름
        try:
            write_xl.save(excelUrl[0])
            processStatusLabel.setText("프로그램이 종료. - 정상")
        except Exception as e:
            print(e)


# class SaveData(QDialog):
#     def __init__(self):
#         super(SaveData, self).__init__()
#         self.saveData()
#
#     def saveData(self):
#         self.setWindowFlags(Qt.Popup)
#         self.setFixedSize(580, 200)
#         self.setStyleSheet("background-color: black;"
#                            "color: white;"
#                            "border: 1px solid yellow;")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    myWindow = StockStart()
    myWindow.show()
    app.exec_()